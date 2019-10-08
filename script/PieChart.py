# -*- coding: utf-8 -*-

import arcpy
import openpyxl,openpyxl.chart,openpyxl.utils
import xlrd,xlwt,os

arcpy.env.overwriteOutput = True
try:
    os.makedirs("C:/yz_spj/temp/excel2txt")
except:
    pass

try:
    arcpy.CheckOutExtension("3D")
except:
    try:
        with open("C:/yz_spj/temp/excel2txt/PieChart3DlicenseERR.txt","w") as err3d:
            err3d.write("License 3D Checkout Field")
    except:
        pass
try:
    arcpy.CheckOutExtension("Spatial")
except:
    try:
        with open("C:/yz_spj/temp/excel2txt/PieChartSpalicenseERR.txt","w") as err3d:
            err3d.write("License Spatial Checkout Field")
    except:
        pass

#Pie Chart Example          F:/1/shp483.shp
# shp输入shp数据，必须具有"XZQNAME", "GRIDCODE"， "mj_sta"字段 注意！！！ "XZQNAME", "GRIDCODE"必须全大写！！
# dissolveFields以列表形式传入要进行融合的字段
# outxlsx 输入全路径形式的xlsx文件名（如果已存在，则会覆盖）#outxlsx = "C:/yz_spj/temp/excel2txt/result.xlsx"
def xls2xlsxWithPieChart(shp, outputpath = "C:/yz_spj/temp/excel2txt", outputname = "result"):
    try:
        os.makedirs("C:/yz_spj/temp/excel2txt")
    except:
        pass

    #判断shp字段并消除大小写问题
    fieldList = arcpy.ListFields(shp)
    for i in fieldList:
        # print i.name
        if "GRIDC" in i.name.upper():
            gridcode = i.name.upper().encode("utf-8")
        elif "XZQNAME" in i.name.upper():
            xzqname = i.name.upper().encode("utf-8")
        elif "MJ_STA" in i.name.upper():
            mj_sta = i.name.lower().encode("utf-8")
        else:
            pass
        print i.name
    list2 = []
    list2.append(gridcode)
    list2.append(xzqname)

    list3 = []
    list3.append(gridcode)

    list4 = []
    list4.append(xzqname)

    outputs = os.path.join(outputpath, outputname + ".xlsx")

    out = arcpy.Dissolve_management(shp, "C:/yz_spj/temp/excel2txt/temp.shp", list2,
                                    [[mj_sta, "SUM"]])

    arcpy.TableToExcel_conversion(out, "C:/yz_spj/temp/excel2txt/excel1te1.xls")  # 与行政区相交后的属性表导出



    wb = xlrd.open_workbook("C:/yz_spj/temp/excel2txt/excel1te1.xls")
    ws = wb.sheet_by_index(0)
    row = ws.row(0)
    for i in row:
        if "GRID" in i.value.upper().encode("utf-8"):   #捕捉"GRIDCODE"列
            griCol = row.index(i)   #列的数字序号
            strgCol = openpyxl.utils.get_column_letter(griCol+1)   #列的英文字符号
            print strgCol
        elif "XZQN" in i.value.upper().encode("utf-8"):
            xzqCol = row.index(i)
            strzCol = openpyxl.utils.get_column_letter(xzqCol+1)
            print strzCol
        elif "mj_s" in i.value.lower().encode("utf-8"):
            mjCol = row.index(i)
            strmCol = openpyxl.utils.get_column_letter(mjCol+1)
            print strmCol
        else:
            pass
    gCol = ws.col(griCol)
    zCol = ws.col(xzqCol)
    mCol = ws.col(mjCol)

    # 新建workbook并逐列写入数据

    new_workbook = xlwt.Workbook()
    worksheet = new_workbook.add_sheet("temp", cell_overwrite_ok=True)
    n = 0
    for i in gCol:
        worksheet.write(n,0,i.value)
        n += 1
    n = 0
    for i in zCol:
        worksheet.write(n,1,i.value)
        n += 1
    n = 0
    for i in mCol:
        worksheet.write(n,2,i.value)
        n += 1

    new_workbook.save("C:/yz_spj/temp/excel2txt/outxls.xls")

    #"XZQNAME" "mj_sta" "GRIDCODE"
    excel1 = xlrd.open_workbook("C:/yz_spj/temp/excel2txt/outxls.xls")
    sheet1 = excel1.sheet_by_index(0)
    nrows = sheet1.nrows
    try:
        os.makedirs("C:/yz_spj/temp/excel2txt")
    except:
        pass
    with open("C:/yz_spj/temp/excel2txt/EXCEL2TXT.txt", "w") as f:
        f.writelines("GRIDCODE,mj_sta,XZQNAME"+"\n")
        for i in range(1, nrows):
            listrow = sheet1.row_values(i)
            listrow2 = listrow[1:2]
            listrow4 = []
            listrow1 = listrow[:1] + listrow[2:]
            for j in listrow2:
                newstr = j.encode("utf-8")
                listrow4.append(newstr)
            str1 = str(listrow1)
            str2 = str1.replace("[", "").replace("]", "").replace(" ", "")
            f.writelines(str2 +"," + listrow4[0] + "\n")
    in2 = "C:/yz_spj/temp/excel2txt/EXCEL2TXT.txt"

    ## 将txt内容写入新xlsx
    newxlsx = "C:/yz_spj/temp/excel2txt/outxlsx1.xlsx"
    #建立新的xlsx文件
    newwb = openpyxl.Workbook()

    #激活sheet
    newws = newwb.active
    n = 1
    with open("C:/yz_spj/temp/excel2txt/EXCEL2TXT.txt", "r") as f:
        for lines in f.readlines():
            templist = lines.split(",")
            if n >= 2:
                ff = float(templist[1])
            else:
                ff = templist[1]
            newws["C{}".format(n)] = ff
            newws["A{}".format(n)] = templist[2]
            newws["B{}".format(n)] = templist[0]
            n += 1

    newwb.save("C:/yz_spj/temp/excel2txt/outxlsx.xlsx")

# 制作饼状图
    wb = openpyxl.load_workbook("C:/yz_spj/temp/excel2txt/outxlsx.xlsx")
    ws1 = wb["Sheet"]
    ws = wb.copy_worksheet(ws1)
    ws["A1"] = u"行政区名称"
    ws["B1"] = u"栅格单元值"
    ws["C1"] = u"区域面积"
    # ws = wb["SHT1"]
    ws.title = u"各行政区各级别面积占比表"
    pie = openpyxl.chart.PieChart()
    nrows = ws.max_row
    labels = openpyxl.chart.Reference(ws, min_col=1, max_col=2, min_row=2, max_row=nrows)
    data = openpyxl.chart.Reference(ws, min_col=3, min_row=2, max_row=nrows)
    pie.title = u"行政区及各级别面积占比图"
    pie.add_data(data)
    pie.set_categories(labels)
    ws.add_chart(pie, "E5")
    #wb.remove_sheet(ws1)
    wb.save(outputs)
    try:
        os.remove("C:/yz_spj/temp/excel2txt/EXCEL2TXT.txt")
    except:
        pass
    try:
        arcpy.Delete_management("C:/yz_spj/temp/excel2txt/temp.shp")
    except:
        pass

    try:
        os.remove("C:/yz_spj/temp/excel2txt/temp.xls")
    except:
        pass
    try:
        os.remove("C:/yz_spj/temp/excel2txt/outxls.xls")
    except:
        pass
    try:
        os.remove("C:/yz_spj/temp/excel2txt/outxlsx1.xlsx")
    except:
        pass

    '''
    xzqname dissolve
    '''

    out = arcpy.Dissolve_management(shp, "C:/yz_spj/temp/excel2txt/temp.shp", list4,
                                    [[mj_sta, "SUM"], [gridcode, "MEAN"]])

    arcpy.TableToExcel_conversion(out, "C:/yz_spj/temp/excel2txt/excel1te1.xls")  # 与行政区相交后的属性表导出

    wb = xlrd.open_workbook("C:/yz_spj/temp/excel2txt/excel1te1.xls")
    ws = wb.sheet_by_index(0)
    row = ws.row(0)
    for i in row:
        if "GRID" in i.value.upper().encode("utf-8"):  # 捕捉"GRIDCODE"列
            griCol = row.index(i)  # 列的数字序号
            strgCol = openpyxl.utils.get_column_letter(griCol + 1)  # 列的英文字符号
#            print strgCol
        elif "XZQN" in i.value.upper().encode("utf-8"):
            xzqCol = row.index(i)
            strzCol = openpyxl.utils.get_column_letter(xzqCol + 1)
#            print strzCol
        elif "mj_s" in i.value.lower().encode("utf-8"):
            mjCol = row.index(i)
            strmCol = openpyxl.utils.get_column_letter(mjCol + 1)
#            print strmCol
        else:
            pass
    gCol = ws.col(griCol)
    zCol = ws.col(xzqCol)
    mCol = ws.col(mjCol)

    # 新建workbook并逐列写入数据

    new_workbook = xlwt.Workbook()
    worksheet = new_workbook.add_sheet("temp", cell_overwrite_ok=True)
    n = 0
    for i in gCol:
        worksheet.write(n, 0, i.value)
        n += 1
    n = 0
    for i in zCol:
        worksheet.write(n, 1, i.value)
        n += 1
    n = 0
    for i in mCol:
        worksheet.write(n, 2, i.value)
        n += 1

    new_workbook.save("C:/yz_spj/temp/excel2txt/outxls.xls")

    # "XZQNAME" "mj_sta" "GRIDCODE"
    excel1 = xlrd.open_workbook("C:/yz_spj/temp/excel2txt/outxls.xls")
    sheet1 = excel1.sheet_by_index(0)
    nrows = sheet1.nrows

    with open("C:/yz_spj/temp/excel2txt/EXCEL2TXT.txt", "w") as f:
        f.writelines("GRIDCODE,mj_sta,XZQNAME" + "\n")
        for i in range(1, nrows):
            listrow = sheet1.row_values(i)
            listrow2 = listrow[1:2]
            listrow4 = []
            listrow1 = listrow[:1] + listrow[2:]
            for j in listrow2:
                newstr = j.encode("utf-8")
                listrow4.append(newstr)
            str1 = str(listrow1)
            str2 = str1.replace("[", "").replace("]", "").replace(" ", "")
            f.writelines(str2 + "," + listrow4[0] + "\n")
    in2 = "C:/yz_spj/temp/excel2txt/EXCEL2TXT.txt"

    ## 将txt内容写入新xlsx
    newxlsx = "C:/yz_spj/temp/excel2txt/outxlsx1.xlsx"
    # 建立新的xlsx文件
    newwb = openpyxl.Workbook()

    # 激活sheet
    newws = newwb.active
    n = 1
    with open("C:/yz_spj/temp/excel2txt/EXCEL2TXT.txt", "r") as f:
        for lines in f.readlines():
            templist = lines.split(",")
            if n >= 2:
                ff = float(templist[1])
            else:
                ff = templist[1]
            newws["C{}".format(n)] = ff
            newws["A{}".format(n)] = templist[2]
            newws["B{}".format(n)] = templist[0]
            n += 1

    newwb.save("C:/yz_spj/temp/excel2txt/outxlsx.xlsx")

    wb = openpyxl.load_workbook("C:/yz_spj/temp/excel2txt/outxlsx.xlsx")
    wstemp = wb["Sheet"]
    wscol1 = wstemp["A"]
    wscol2 = wstemp["B"]
    wscol3 = wstemp["C"]



    # 制作饼状图
    wb = openpyxl.load_workbook(outputs)
    ws1 = wb.create_sheet()
    # ws1 = wb["Sheet"]
    ws2 = wb.copy_worksheet(ws1)
    ws2.title = u"各行政区面积占比表"

    n = 1
    for i in wscol1:
        ws2.cell(row=n, column=1).value = i.value
        n += 1
    n = 1
    for i in wscol2:
        ws2.cell(row=n, column=2).value = i.value
        n += 1
        #print ws2.cell(row=n, column=2).value
    n = 1
    for i in wscol3:
        ws2.cell(row=n, column=3).value = i.value
        n += 1

    ws2["A1"] = u"行政区名称"
    ws2["B1"] = u"行政区栅格单元均值"
    ws2["C1"] = u"区域面积"
    # ws = wb["SHT1"]
    pie = openpyxl.chart.PieChart()
    nrows = ws2.max_row
    labels = openpyxl.chart.Reference(ws2, min_col=1, min_row=2, max_row=nrows)
    data = openpyxl.chart.Reference(ws2, min_col=3, min_row=2, max_row=nrows)
    pie.title = u"各行政区面积占比图"
    pie.add_data(data)
    pie.set_categories(labels)
    ws2.add_chart(pie, "E5")
    #wb.remove_sheet(ws1)
    wb.save(outputs)
    try:
        os.remove("C:/yz_spj/temp/excel2txt/EXCEL2TXT.txt")
    except:
        pass
    try:
        arcpy.Delete_management("C:/yz_spj/temp/excel2txt/temp.shp")
    except:
        pass

    try:
        os.remove("C:/yz_spj/temp/excel2txt/temp.xls")
    except:
        pass
    try:
        os.remove("C:/yz_spj/temp/excel2txt/outxls.xls")
    except:
        pass
    try:
        os.remove("C:/yz_spj/temp/excel2txt/outxlsx1.xlsx")
    except:
        pass

    '''
    gridcode dissolve
    '''

    out = arcpy.Dissolve_management(shp, "C:/yz_spj/temp/excel2txt/temp.shp", list3,
                                    [[mj_sta, "SUM"], [xzqname, "FIRST"]])

    arcpy.TableToExcel_conversion(out, "C:/yz_spj/temp/excel2txt/excel1te1.xls")  # 与行政区相交后的属性表导出

    wb = xlrd.open_workbook("C:/yz_spj/temp/excel2txt/excel1te1.xls")
    ws = wb.sheet_by_index(0)
    row = ws.row(0)
    for i in row:
        if "GRID" in i.value.upper().encode("utf-8"):  # 捕捉"GRIDCODE"列
            griCol = row.index(i)  # 列的数字序号
            strgCol = openpyxl.utils.get_column_letter(griCol + 1)  # 列的英文字符号
#            print strgCol
        elif "XZQN" in i.value.upper().encode("utf-8"):
            xzqCol = row.index(i)
            strzCol = openpyxl.utils.get_column_letter(xzqCol + 1)
#            print strzCol
        elif "mj_s" in i.value.lower().encode("utf-8"):
            mjCol = row.index(i)
            strmCol = openpyxl.utils.get_column_letter(mjCol + 1)
#            print strmCol
        else:
            pass
    gCol = ws.col(griCol)
    zCol = ws.col(xzqCol)
    mCol = ws.col(mjCol)

    # 新建workbook并逐列写入数据

    new_workbook = xlwt.Workbook()
    worksheet = new_workbook.add_sheet("temp", cell_overwrite_ok=True)
    n = 0
    for i in gCol:
        worksheet.write(n, 0, i.value)
        n += 1
    n = 0
    for i in zCol:
        worksheet.write(n, 1, i.value)
        n += 1
    n = 0
    for i in mCol:
        worksheet.write(n, 2, i.value)
        n += 1

    new_workbook.save("C:/yz_spj/temp/excel2txt/outxls.xls")

    # "XZQNAME" "mj_sta" "GRIDCODE"
    excel1 = xlrd.open_workbook("C:/yz_spj/temp/excel2txt/outxls.xls")
    sheet1 = excel1.sheet_by_index(0)
    nrows = sheet1.nrows

    with open("C:/yz_spj/temp/excel2txt/EXCEL2TXT.txt", "w") as f:
        f.writelines("GRIDCODE,mj_sta,XZQNAME" + "\n")
        for i in range(1, nrows):
            listrow = sheet1.row_values(i)
            listrow2 = listrow[1:2]
            listrow4 = []
            listrow1 = listrow[:1] + listrow[2:]
            for j in listrow2:
                newstr = j.encode("utf-8")
                listrow4.append(newstr)
            str1 = str(listrow1)
            str2 = str1.replace("[", "").replace("]", "").replace(" ", "")
            f.writelines(str2 + "," + listrow4[0] + "\n")
    in2 = "C:/yz_spj/temp/excel2txt/EXCEL2TXT.txt"

    ## 将txt内容写入新xlsx
    newxlsx = "C:/yz_spj/temp/excel2txt/outxlsx1.xlsx"
    # 建立新的xlsx文件
    newwb = openpyxl.Workbook()

    # 激活sheet
    newws = newwb.active
    n = 1
    with open("C:/yz_spj/temp/excel2txt/EXCEL2TXT.txt", "r") as f:
        for lines in f.readlines():
            templist = lines.split(",")
            if n >= 2:
                ff = float(templist[1])
            else:
                ff = templist[1]
            newws["C{}".format(n)] = ff
            newws["A{}".format(n)] = templist[2]
            newws["B{}".format(n)] = templist[0]
            n += 1

    newwb.save("C:/yz_spj/temp/excel2txt/outxlsx.xlsx")

    wb = openpyxl.load_workbook("C:/yz_spj/temp/excel2txt/outxlsx.xlsx")
    wstemp = wb["Sheet"]
    wscol1 = wstemp["A"]
    wscol2 = wstemp["B"]
    wscol3 = wstemp["C"]



    # 制作饼状图
    wb = openpyxl.load_workbook(outputs)
    ws1 = wb.create_sheet()
    # ws1 = wb["Sheet"]
    ws3 = wb.copy_worksheet(ws1)
    ws3.title = u"各级别面积占比表"

    # n = 1
    # for i in wscol1:
    #     ws3.cell(row=n, column=1).value = i.value
    #     n += 1
    n = 1
    for i in wscol2:
        ws3.cell(row=n, column=1).value = i.value
        n += 1
        #print ws2.cell(row=n, column=2).value
    n = 1
    for i in wscol3:
        ws3.cell(row=n, column=2).value = i.value
        n += 1

    # ws3["A1"] = u"行政区名称"
    ws3["A1"] = u"栅格单元值"
    ws3["B1"] = u"区域面积"
    # ws = wb["SHT1"]
    pie = openpyxl.chart.PieChart()
    nrows = ws3.max_row
    labels = openpyxl.chart.Reference(ws3, min_col=1, min_row=2, max_row=nrows)
    data = openpyxl.chart.Reference(ws3, min_col=2, min_row=2, max_row=nrows)
    pie.title = u"各级别面积占比图"
    pie.add_data(data)
    pie.set_categories(labels)
    ws3.add_chart(pie, "E5")
    del wb["Sheet"]
    del wb["Sheet1"]
    del wb["Sheet2"]
    wb.save(outputs)
    try:
        os.remove("C:/yz_spj/temp/excel2txt/EXCEL2TXT.txt")
    except:
        pass
    try:
        arcpy.Delete_management("C:/yz_spj/temp/excel2txt/temp.shp")
    except:
        pass

    try:
        os.remove("C:/yz_spj/temp/excel2txt/temp.xls")
    except:
        pass
    try:
        os.remove("C:/yz_spj/temp/excel2txt/outxls.xls")
    except:
        pass
    try:
        os.remove("C:/yz_spj/temp/excel2txt/outxlsx.xlsx")
    except:
        pass


#
# shp = "D:/1/kaer11.shp"         #矢量数据融合数据
# outpath = "D:/1/22"
# outname = "kaer444"
# excels = "C:/yz_spj/temp/excel2txt/temp.xls"
# outxlsx = "F:/1/11/outexcel8.xlsx"
# dissolveFields = ["XZQNAME"]
#["XZQNAME", "GRIDCODE"]
# shp = "D:/1/33/ggggggggg_XZQ.shp"
# #
# xls2xlsxWithPieChart(shp, outpath, outname)









